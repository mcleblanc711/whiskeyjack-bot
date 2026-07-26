#!/usr/bin/env python3
"""Build the backlog workbook from the four CSVs in docs/backlog/.

The CSVs are authoritative; the workbook is a build output and is not tracked. It
used to be tracked and hand-patched, which drifted from the CSVs at least once
(PR #13) and made the Status flip a two-file edit -- the step that was missed in
M1-203, M1-401 and M1-305.

    uv run python scripts/backlog_xlsx.py            # writes ./whiskeyjack-bot-v1-backlog.xlsx
    uv run python scripts/backlog_xlsx.py --out /tmp/backlog.xlsx

Output is deterministic: sheets, row order and column order all follow the CSVs.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT: Final = Path(__file__).resolve().parents[1]
BACKLOG_DIR: Final = REPO_ROOT / "docs" / "backlog"
DEFAULT_OUT: Final = REPO_ROOT / "whiskeyjack-bot-v1-backlog.xlsx"

# (sheet name, csv file, excel table name). Order matches the tracked workbook this
# replaces, so an existing bookmark or reference still lands on the same tab.
SHEETS: Final = (
    ("Backlog", "backlog.csv", "BacklogTable"),
    ("Verified Facts", "verified-facts.csv", "VerifiedFactsTable"),
    ("Decisions", "decisions.csv", "DecisionsTable"),
    ("Risks", "risks.csv", "RisksTable"),
)

MIN_WIDTH: Final = 10
MAX_WIDTH: Final = 60


def _read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = [record for record in csv.reader(handle) if record]
    if len(rows) < 2:
        raise SystemExit(f"FAIL: {path} has no data rows.")
    width = len(rows[0])
    for offset, record in enumerate(rows[1:], start=2):
        if len(record) != width:
            raise SystemExit(
                f"FAIL: {path} line {offset} has {len(record)} fields, expected {width}."
            )
    return rows


def _fill_sheet(sheet: Worksheet, rows: list[list[str]], table_name: str) -> None:
    for record in rows:
        sheet.append(record)

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="top")

    for column_index in range(1, len(rows[0]) + 1):
        letter = get_column_letter(column_index)
        longest = max(len(record[column_index - 1]) for record in rows)
        sheet.column_dimensions[letter].width = min(MAX_WIDTH, max(MIN_WIDTH, longest + 2))
        for cell in sheet[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.font = header_font

    sheet.freeze_panes = "A2"
    reference = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
    table = Table(displayName=table_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help=f"output path (default: {DEFAULT_OUT.relative_to(REPO_ROOT)})",
    )
    args = parser.parse_args(argv)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, csv_name, table_name in SHEETS:
        rows = _read_csv(BACKLOG_DIR / csv_name)
        sheet = workbook.create_sheet(title=sheet_name)
        _fill_sheet(sheet, rows, table_name)
        print(f"{sheet_name}: {len(rows) - 1} rows from docs/backlog/{csv_name}")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
